"""Structured data store: accounts, orders, tickets from the supplied workbook."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from typing import Any

import openpyxl

from app.config import TIMEZONE, WORKBOOK


def _dt(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=TIMEZONE) if value.tzinfo is None else value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=TIMEZONE)
        except ValueError:
            continue
    return None


def _bool(value: Any) -> bool | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "yes", "1", "y")


@dataclass
class Account:
    account_id: str
    account_name: str
    plan: str
    status: str
    csm: str | None
    contract_file: str | None
    premium_support: bool | None
    notes: str | None

    @property
    def has_contract(self) -> bool:
        return bool(self.contract_file)


@dataclass
class Order:
    order_id: str
    account_id: str
    carrier: str
    status: str
    booked_at: datetime | None
    pickup_window_start: datetime | None
    pickup_window_end: datetime | None
    pickup_actual_at: datetime | None
    shipment_fee_inr: float | None
    carrier_fault: bool | None
    customer_fault: bool | None
    cancellation_requested_at: datetime | None
    notes: str | None


@dataclass
class Ticket:
    ticket_id: str
    account_id: str
    created_at: datetime | None
    status: str
    subject: str
    description: str
    channel: str | None
    assigned_to: str | None
    last_customer_message_at: datetime | None
    historical_resolution: str | None

    @property
    def is_open(self) -> bool:
        return (self.status or "").lower() == "open"

    @property
    def customer_followed_up(self) -> bool:
        return bool(
            self.last_customer_message_at
            and self.created_at
            and self.last_customer_message_at > self.created_at
        )


@dataclass
class Snapshot:
    """Dataset metadata from the README sheet."""

    taken_at: datetime
    raw: str
    currency: str
    notes: list[str] = field(default_factory=list)


def _rows(ws) -> list[dict]:
    it = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(it)]
    except StopIteration:
        return []
    out = []
    for row in it:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return out


@dataclass
class Store:
    snapshot: Snapshot
    accounts: dict[str, Account]
    orders: dict[str, Order]
    tickets: dict[str, Ticket]

    def account(self, account_id: str) -> Account | None:
        return self.accounts.get(account_id)

    def account_by_name(self, name: str) -> Account | None:
        want = re.sub(r"[^a-z0-9]", "", (name or "").lower())
        if not want:
            return None
        for acc in self.accounts.values():
            have = re.sub(r"[^a-z0-9]", "", acc.account_name.lower())
            if have == want or want in have or have in want:
                return acc
        return None

    def orders_for(self, account_id: str) -> list[Order]:
        return [o for o in self.orders.values() if o.account_id == account_id]

    def tickets_for(self, account_id: str) -> list[Ticket]:
        return [t for t in self.tickets.values() if t.account_id == account_id]

    @property
    def now(self) -> datetime:
        # The workbook snapshot time is the only "now" the system uses.
        return self.snapshot.taken_at


@lru_cache(maxsize=1)
def load_store() -> Store:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    readme = {}
    for row in wb["README"].iter_rows(values_only=True):
        if row and row[0]:
            readme.setdefault(str(row[0]).strip(), str(row[1]).strip() if len(row) > 1 and row[1] else "")

    raw_snap = readme.get("Dataset snapshot", "")
    m = re.match(r"(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2})", raw_snap)
    if not m:
        raise RuntimeError(f"Cannot parse dataset snapshot from README: {raw_snap!r}")
    taken_at = datetime.strptime(m.group(1).replace("T", " "), "%Y-%m-%d %H:%M").replace(tzinfo=TIMEZONE)

    snapshot = Snapshot(
        taken_at=taken_at,
        raw=raw_snap,
        currency=readme.get("Currency", "INR"),
        notes=[v for k, v in readme.items() if k in ("Notes", "Important") and v],
    )

    accounts = {
        r["account_id"]: Account(
            account_id=r["account_id"],
            account_name=r["account_name"],
            plan=r["plan"],
            status=r["status"],
            csm=r.get("csm") or None,
            contract_file=r.get("contract_file") or None,
            premium_support=_bool(r.get("premium_support")),
            notes=r.get("notes") or None,
        )
        for r in _rows(wb["accounts"])
    }

    orders = {
        r["order_id"]: Order(
            order_id=r["order_id"],
            account_id=r["account_id"],
            carrier=r.get("carrier") or "",
            status=(r.get("status") or "").upper(),
            booked_at=_dt(r.get("booked_at")),
            pickup_window_start=_dt(r.get("pickup_window_start")),
            pickup_window_end=_dt(r.get("pickup_window_end")),
            pickup_actual_at=_dt(r.get("pickup_actual_at")),
            shipment_fee_inr=float(r["shipment_fee_inr"]) if r.get("shipment_fee_inr") else None,
            carrier_fault=_bool(r.get("carrier_fault")),
            customer_fault=_bool(r.get("customer_fault")),
            cancellation_requested_at=_dt(r.get("cancellation_requested_at")),
            notes=r.get("notes") or None,
        )
        for r in _rows(wb["orders"])
    }

    tickets = {
        r["ticket_id"]: Ticket(
            ticket_id=r["ticket_id"],
            account_id=r["account_id"],
            created_at=_dt(r.get("created_at")),
            status=(r.get("status") or "").lower(),
            subject=r.get("subject") or "",
            description=r.get("description") or "",
            channel=r.get("channel") or None,
            assigned_to=r.get("assigned_to") or None,
            last_customer_message_at=_dt(r.get("last_customer_message_at")),
            historical_resolution=r.get("historical_resolution") or None,
        )
        for r in _rows(wb["tickets"])
    }

    return Store(snapshot=snapshot, accounts=accounts, orders=orders, tickets=tickets)
